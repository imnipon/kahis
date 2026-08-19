#!/usr/bin/env python3
"""Generate update_clean.csv, update_errors.csv, update_summary.txt from XLS source."""
import csv
import openpyxl
from collections import defaultdict
from pathlib import Path

MASTER_CSV = Path(r"C:\Users\C18\Documents\Cursor\kahis\resource\original\items_export_enabled_AT_20260817.csv")
XLS_FILE = Path(r"C:\Users\C18\Documents\Cursor\kahis\resource\original\order_tx-2026-07-24.xlsx")
OUT_DIR = Path(r"C:\Users\C18\Documents\Cursor\kahis\resource\updateitemsfiles")

# --- Load master item codes (strip leading apostrophe) ---
master_codes = set()
with open(MASTER_CSV, encoding="utf-8-sig", newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        code = row.get("Item Code(รหัส)") or row.get("Item Code") or ""
        code = str(code).strip().lstrip("'").strip()
        if code:
            master_codes.add(code)

print(f"Master codes loaded: {len(master_codes)}")

# --- Read XLS ---
wb = openpyxl.load_workbook(str(XLS_FILE), read_only=True)
all_sheets = wb.sheetnames

# Skip the summary/export sheet (first sheet which has a header row describing itself)
SKIP_SHEETS = {"order_tx-2026-07-17020734"}
TAB_SHEETS = [s for s in all_sheets if s not in SKIP_SHEETS]

# Col indices (0-based): item_code=3, description=4, route=8, frequency=9, note=10
COL_ITEM_CODE = 3
COL_DESC = 4
COL_ROUTE = 8
COL_FREQ = 9
COL_NOTE = 10

def normalize(v):
    if v is None:
        return ""
    return str(v).strip()

errors = []  # list of dicts
# placements: item_code -> list of {sheet, route, freq, note, desc}
placements = defaultdict(list)

total_placements = 0
sheet_row_counts = {}

for sheet_name in TAB_SHEETS:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Detect if first row is a header (contains 'Route' or None in col 3)
    start_idx = 0
    if rows and rows[0][COL_ITEM_CODE] is None and rows[0][COL_ROUTE] is not None:
        start_idx = 1  # skip header row
    
    data_rows = rows[start_idx:]
    sheet_seen_codes = defaultdict(list)
    sheet_count = 0
    
    for row in data_rows:
        # Skip completely empty rows
        if all(c is None for c in row):
            continue
        
        item_code_raw = row[COL_ITEM_CODE] if len(row) > COL_ITEM_CODE else None
        desc = normalize(row[COL_DESC] if len(row) > COL_DESC else None)
        route = normalize(row[COL_ROUTE] if len(row) > COL_ROUTE else None)
        freq = normalize(row[COL_FREQ] if len(row) > COL_FREQ else None)
        note = normalize(row[COL_NOTE] if len(row) > COL_NOTE else None)
        
        # EMPTY_CODE
        if item_code_raw is None or str(item_code_raw).strip() == "":
            errors.append({
                "Item Code": "",
                "Sheet": sheet_name,
                "Description (from XLS)": desc,
                "Error Type": "EMPTY_CODE",
                "Detail": "Row has no Item Code"
            })
            continue
        
        # Normalize: strip apostrophe, then zero-pad to 12 digits to match master
        raw_str = str(item_code_raw).strip().lstrip("'").strip()
        try:
            item_code = str(int(float(raw_str))).zfill(12)
        except (ValueError, TypeError):
            item_code = raw_str.zfill(12) if raw_str.isdigit() else raw_str
        total_placements += 1
        sheet_count += 1
        
        # NOT_IN_MASTER
        if item_code not in master_codes:
            errors.append({
                "Item Code": item_code,
                "Sheet": sheet_name,
                "Description (from XLS)": desc,
                "Error Type": "NOT_IN_MASTER",
                "Detail": f"Code {item_code} not found in master CSV"
            })
            continue
        
        sheet_seen_codes[item_code].append({"sheet": sheet_name, "route": route, "freq": freq, "note": note, "desc": desc})
        placements[item_code].append({"sheet": sheet_name, "route": route, "freq": freq, "note": note, "desc": desc})
    
    # DUPLICATE_IN_SHEET
    for code, occurrences in sheet_seen_codes.items():
        if len(occurrences) > 1:
            for occ in occurrences[1:]:  # report extras
                errors.append({
                    "Item Code": code,
                    "Sheet": sheet_name,
                    "Description (from XLS)": occ["desc"],
                    "Error Type": "DUPLICATE_IN_SHEET",
                    "Detail": f"Code {code} appears {len(occurrences)} times in sheet {sheet_name}"
                })
                # Remove duplicate from placements (keep only first)
            # Keep only first occurrence
            placements[code] = [p for p in placements[code] if not (p["sheet"] == sheet_name)] + [occurrences[0]]
    
    sheet_row_counts[sheet_name] = sheet_count

wb.close()

# --- Build update_clean.csv ---
clean_rows = []
single_placement = 0
multi_same = 0
multi_conflict = 0
conflict_items = []

for item_code, plist in sorted(placements.items()):
    if len(plist) == 1:
        p = plist[0]
        tag = p["sheet"]
        route, freq, note = p["route"], p["freq"], p["note"]
        # Skip if all Tx fields empty
        if not any([tag, route, freq, note]):
            continue
        clean_rows.append({
            "Item Code": item_code,
            "Base Code": "",
            "Tag": tag,
            "Route": route,
            "Dose": "",  # Dose not separately identified; freq col may contain dose/rate
            "Frequency": freq,
            "Note": note
        })
        single_placement += 1
    else:
        # Check if all have identical Route+Freq+Note
        def tx_key(p):
            return (p["route"], p["freq"], p["note"])
        
        keys = [tx_key(p) for p in plist]
        if len(set(keys)) == 1:
            # All same Tx
            tags = "|".join(p["sheet"] for p in plist)
            p = plist[0]
            if not any([tags, p["route"], p["freq"], p["note"]]):
                continue
            clean_rows.append({
                "Item Code": item_code,
                "Base Code": "",
                "Tag": tags,
                "Route": p["route"],
                "Dose": "",
                "Frequency": p["freq"],
                "Note": p["note"]
            })
            multi_same += 1
        else:
            # Conflict - sub-codes
            multi_conflict += 1
            conflict_items.append(f"{item_code} ({', '.join(p['sheet'] for p in plist)})")
            # First placement: item_code, Base Code empty
            p0 = plist[0]
            clean_rows.append({
                "Item Code": item_code,
                "Base Code": "",
                "Tag": p0["sheet"],
                "Route": p0["route"],
                "Dose": "",
                "Frequency": p0["freq"],
                "Note": p0["note"]
            })
            # Rest: ITEMCODE-1, ITEMCODE-2, ...
            for idx, p in enumerate(plist[1:], start=1):
                clean_rows.append({
                    "Item Code": f"{item_code}-{idx}",
                    "Base Code": item_code,
                    "Tag": p["sheet"],
                    "Route": p["route"],
                    "Dose": "",
                    "Frequency": p["freq"],
                    "Note": p["note"]
                })

# Write update_clean.csv
clean_path = OUT_DIR / "update_clean.csv"
with open(clean_path, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["Item Code", "Base Code", "Tag", "Route", "Dose", "Frequency", "Note"])
    writer.writeheader()
    writer.writerows(clean_rows)

# Write update_errors.csv
errors_path = OUT_DIR / "update_errors.csv"
with open(errors_path, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["Item Code", "Sheet", "Description (from XLS)", "Error Type", "Detail"])
    writer.writeheader()
    writer.writerows(errors)

# Count errors by type
error_counts = defaultdict(int)
for e in errors:
    error_counts[e["Error Type"]] += 1

valid_placements = sum(len(v) for v in placements.values())

# Write update_summary.txt
summary_path = OUT_DIR / "update_summary.txt"
with open(summary_path, "w", encoding="utf-8") as f:
    f.write("=== Update Summary ===\n\n")
    f.write(f"Total XLS placements read: {total_placements}\n\n")
    f.write("Sheets (valid rows read):\n")
    for s, cnt in sheet_row_counts.items():
        f.write(f"  {s}: {cnt}\n")
    f.write(f"\nValid placements (code in master): {valid_placements}\n")
    f.write(f"Error rows: {sum(error_counts.values())}\n")
    for etype, cnt in error_counts.items():
        f.write(f"  {etype}: {cnt}\n")
    f.write(f"\nSingle-placement items: {single_placement}\n")
    f.write(f"Multi-placement same Tx: {multi_same}\n")
    f.write(f"Multi-placement conflict Tx (sub-codes): {multi_conflict}\n")
    if conflict_items:
        f.write("\nConflict item codes:\n")
        for ci in conflict_items:
            f.write(f"  {ci}\n")

print("\n=== Summary ===")
print(f"Total XLS placements read: {total_placements}")
print(f"Sheets: {sheet_row_counts}")
print(f"Valid placements: {valid_placements}")
print(f"Errors: {dict(error_counts)}")
print(f"Single-placement: {single_placement}")
print(f"Multi-placement same Tx: {multi_same}")
print(f"Multi-placement conflict Tx: {multi_conflict}")
print(f"Conflict items: {conflict_items}")
print(f"\nFiles written:")
print(f"  {clean_path}")
print(f"  {errors_path}")
print(f"  {summary_path}")
